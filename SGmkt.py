# -*- coding:utf-8 -*-

def main(date, cookie, location1, location2, location3):
    try:

        import requests, json, datetime

        # date = str(datetime.datetime.now().date())
        # date = '2022/06/14'   # 测试用，指定数据时间
        # print(date)
        date = date.replace('-', '/')

        url = 'http://xuri.p4p.sogou.com/statistic/showPlanStatisticData.action'

        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36'
        }

        data = {
            'planStatus': '-1',
            'planMatchType': '0',
            'planSearchInfo': '',
            'timeType': '6',
            'filterType': '0',
            'marketTargetType': '0',
            'packageType': '0',
            'startDate': 'yyyy-mm-dd',
            'endDate': 'yyyy-mm-dd'
        }

        # 时间'yyyy-mm-dd'替换文当天时间date
        data['startDate'] = data['startDate'].replace("yyyy-mm-dd", date)
        data['endDate'] = data['endDate'].replace("yyyy-mm-dd", date)

        response = requests.post(url, headers=headers, cookies=cookie, data=data)

        result = response.text
        # result = json.loads(result)
        # json_res = json.loads(response.text.encode('utf-8'))

        # print(result)  # 打印直接爬虫结果，debug时使用

        result = response.text
        result = result.replace('[]', '""')
        result = result.replace('[', '')
        result = result.replace(']', '')

        result = json.loads(result)

        # print(result)   # 打印字符处理后的爬虫结果，debug时使用

        impression = result['data']['showpv'].replace('--', '0')
        click = result['data']['showclick']
        cost = result['data']['showconsume']
       
        print(impression, click, cost)  # 打印展点消

    except:
        impression = 'error'
        click = 'error'
        cost = 'error'



    # 写入Excel

    import openpyxl

    # 导入模板
    templ = r'推广数据.xlsx'
    wb = openpyxl.load_workbook(templ)
    ws = wb['Sheet1']

    # 写入数据
    ws[location1].value = impression
    ws[location2].value = click
    ws[location3].value = cost

    # 保存文件  过程中需关闭文件，执行完成后打开即可
    wb.save(r'推广数据.xlsx')
