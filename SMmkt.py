# -*- coding:utf-8 -*-


import requests, json, re, time, datetime

def main(date, cookie, headers, smid, location1, location2, location3):

    try:

        # date = str(datetime.datetime.now().date())
        # date = '2022/06/14'   # 测试用，指定数据时间
        # print(date)
        date = date.replace('-', '')

        url = 'https://e.uc.cn/shc/web/main/report/shc/stat?userId=smid'
        url = url.replace('smid', smid)

        # 重点：network中想要的请求，没有Formdata，只有Payload时，这里作为身份识别，requests中的data使用这里！！！
        payload_data = {"aggregate": 3, "filterConditions": [], "dateParam": {"start": "yyyy-mm-dd", "end": "yyyy-mm-dd"},
                        "adReportType": 1,
                        "overview": 'false', "download": 'false', "chart": 'false',
                        "pageParam": {"pageSize": 20, "pageNo": 1, "sort": "", "asc": 'false'}}

        # 时间'abc'替换文当天时间today
        payload_data['dateParam']['start'] = payload_data['dateParam']['start'].replace("yyyy-mm-dd", date)
        payload_data['dateParam']['end'] = payload_data['dateParam']['end'].replace("yyyy-mm-dd", date)


        response = requests.post(url, headers=headers, cookies=cookie, data=json.dumps(payload_data))  # 将payload_data通过json.dumps()方法编码为json数据

        # print(response)  # 返回网页请求状态，200为正常

        result = response.text
        # result = json.loads(result)
        # json_res = json.loads(response.text.encode('utf-8'))

        # print(result)  # 打印直接爬虫结果，debug时使用

        result = response.text
        result = result.replace('[]', '""')
        result = result.replace('[', '')
        result = result.replace(']', '')

        result = json.loads(result)


        # print(result)   # 打印字符处理后的爬虫结果，debug时使用

    except:
        impression = 'error'
        click = 'error'
        cost = 'error'


    impression = result['data']['list']['srch']
    click = result['data']['list']['click']
    cost = result['data']['list']['consume']

    print(impression, click, cost)  # 打印展点消

    
    
    # 写入Excel
    import openpyxl

    # 导入模板
    templ = r'推广数据.xlsx'
    wb = openpyxl.load_workbook(templ)
    ws = wb['Sheet1']

    # 写入数据

    ws[location1].value = impression
    ws[location2].value = click
    ws[location3].value = cost

    # 保存文件  过程中需关闭文件，执行完成后打开即可
    wb.save(r'推广数据.xlsx')
