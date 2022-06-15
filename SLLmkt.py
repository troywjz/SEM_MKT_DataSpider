# -*- coding:utf-8 -*-

def main(date, cookie, location1, location2, location3):

    try:

        import requests, json, datetime

        # today = str(datetime.datetime.now().date())
        # print(today)

        url = 'https://dianjing.e.360.cn/djapi/report/getreport'


        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36'
        }

        data = {
            'baogao_check': '1',
            'ceng_check': '1',
            'dimension_split':'0',
            'end_date': 'yyyy-mm-dd',      # 时间替换为yyyy-mm-dd，后续会自动替换成当天时间today
            'group_day': '0',
            'ps': '100',
            'source_type': '0',
            'start_date': 'yyyy-mm-dd',      # 时间替换为yyyy-mm-dd，后续会自动替换成当天时间today
            'topn_key': '2',
            'topn_limit': '5000',
        }

        # 时间'yyyy-mm-dd'替换文当天时间date
        data['end_date'] = data['end_date'].replace("yyyy-mm-dd", date)
        data['start_date'] = data['start_date'].replace("yyyy-mm-dd", date)

        response = requests.post(url, headers=headers, cookies=cookie, data=data)

        result = response.text
        result = json.loads(result)
        # json_res = json.loads(response.text.encode('utf-8'))

        # print(result)  # 打印直接爬虫结果，debug时使用

        result = response.text
        result = result.replace('[','')
        result = result.replace(']','')

        result = json.loads(result)

        # print(result)   # 打印字符处理后的爬虫结果，debug时使用

        impression = result['list']['views']
        click = result['list']['clicks']
        cost = result['list']['total_cost']
      
    except:
        impression = 'error'
        click = 'error'
        cost = 'error'

    print(impression, click, cost)  # 打印展点消




    # 写入Excel
    import openpyxl

    # 导入模板
    templ = r'推广数据.xlsx'
    wb = openpyxl.load_workbook(templ)
    ws = wb['Sheet1']

    # 写入数据
    ws[location1].value = impression
    ws[location2].value = click
    ws[location3].value = cost

    # 保存文件  过程中需关闭文件，执行完成后打开即可
    wb.save(r'推广数据.xlsx')

